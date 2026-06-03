#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
P&ID管道数据提取核心逻辑
供 pid_extractor.py (CLI) 和 pid_extractor_gui.py (GUI) 共同使用
"""

import re
import os
import sys
import logging
import pandas as pd

logger = logging.getLogger(__name__)

SUPPORTED_PROJECT_TYPES = ['巨化项目', '乌兹项目', '天华项目', '金昱元项目']

PROJECT_FORMAT_EXAMPLES = {
    "巨化项目":  "示例: 4101BRR-02457-200-03CBMB1-H",
    "乌兹项目":  "示例: PA-2001002A-100-C1C-N",
    "天华项目":  "示例: 01PL-216061-125-C22S-H",
    "金昱元项目": "示例: 01CSL03203-150-C1N-H",
}


def get_resource_path(relative_path):
    """获取资源文件路径（支持打包后的 exe）"""
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


def normalize_text(s):
    """文本标准化，清理不可见字符"""
    import unicodedata
    s = str(s).strip()
    s = unicodedata.normalize('NFKC', s)
    s = s.replace('\x00', '')
    s = re.sub(r'[‐-―]', '-', s)
    s = re.sub(r'[\x00-\x1F\x7F-\x9F]', '', s)
    s = re.sub(r'\s+-', '-', s)   # 清理连字符前的空格，如 "50 -C1S" → "50-C1S"
    return s


def load_medium_codes(code_file_path):
    """从 Excel 文件加载介质代码映射"""
    try:
        # keep_default_na=False 防止 "NA" 被 pandas 误读为 NaN
        df = pd.read_excel(code_file_path, header=None, keep_default_na=False)
        medium_codes = {}
        for _, row in df.iterrows():
            code = str(row.iloc[0]).strip()
            name = str(row.iloc[1]).strip()
            if not code or code == 'nan':
                continue
            if not name or name == 'nan':
                continue
            medium_codes[code] = name
        logger.info(f"成功加载 {len(medium_codes)} 个介质代码")
        return medium_codes
    except Exception as e:
        logger.error(f"无法加载介质代码文件: {e}")
        return {}


def determine_phase(medium_name):
    """根据介质名称判断相态"""
    gas_keywords = ['蒸汽', '汽', '气', '空气', '氢气', '氮气', '氧气', '二氧化碳', '天然气', '废气']
    liquid_keywords = ['水', '油', '液', '溶液', '酸', '碱', '汽油', '柴油', '凝结']
    for kw in gas_keywords:
        if kw in medium_name:
            return '气相'
    for kw in liquid_keywords:
        if kw in medium_name:
            return '液相'
    return '未知相态'


def extract_text_from_dwg(dwg_path, log_fn=None):
    """从 DWG 文件中提取文本（需要本地 AutoCAD）"""
    def log(msg):
        if log_fn:
            log_fn(msg)
        else:
            logger.info(msg)

    try:
        from pyautocad import Autocad
        acad = Autocad(create_if_not_exists=True)
        log("成功连接到AutoCAD")

        abs_path = os.path.abspath(dwg_path)
        log(f"打开文件: {abs_path}")
        doc = acad.app.Documents.Open(abs_path)
        log(f"成功打开文件: {doc.Name}")

        model_space = doc.ModelSpace
        log(f"模型空间实体数量: {model_space.Count}")

        text_entities = []
        total = model_space.Count
        for i in range(total):
            try:
                if i % 10000 == 0:
                    log(f"处理进度: {i}/{total} ({i / total * 100:.1f}%)")
                entity = model_space.Item(i)
                entity_type = entity.ObjectName
                if entity_type in ("AcDbText", "AcDbMText", "AcDbBlockReference"):
                    if entity_type in ("AcDbText", "AcDbMText"):
                        content = entity.TextString
                        if content:
                            text_entities.append(content)
                    else:
                        try:
                            if hasattr(entity, 'GetAttributes'):
                                for attr in entity.GetAttributes():
                                    if hasattr(attr, 'TextString'):
                                        text_entities.append(attr.TextString)
                        except Exception:
                            pass
            except Exception:
                continue

        log(f"提取了 {len(text_entities)} 个文本")
        doc.Close(False)
        log("已关闭文档")
        return text_entities
    except Exception as e:
        log(f"提取文本失败: {e}")
        return []


def find_pipeline_numbers(text_entities, project_type, log_fn=None):
    """查找管道号，根据项目类型选择正则模式"""
    def log(msg):
        if log_fn:
            log_fn(msg)
        else:
            logger.info(msg)

    if project_type == "乌兹项目":
        patterns = [
            r'([A-Z0-9]{1,4})-([A-Z0-9]{4,8})-(\d{2,4})-([A-Z0-9]{1,4})-([A-Z0-9]{1,2})',
            r'([A-Z0-9]{1,4})-([A-Z0-9]{4,8})-(\d{2,4})$',
        ]
        test_strings = [
            'PA-2001002A-100-C1C-N',
            'BW-2001003B-150-D2D-H',
            'PA-2001004C-200-E3E-C',
            'PA-2001005D-50',
        ]
    elif project_type == "天华项目":
        patterns = [
            r'(\d{2}[A-Za-z]{2,4})-(\d{5,7}[A-Z]?)-(\d{2,4})-([A-Z0-9]{2,6})-([A-Z]{1,2})',
            r'(\d{2}[A-Za-z]{2,4})-(\d{5,7}[A-Z]?)-(\d{2,4})$',
        ]
        test_strings = [
            '01PL-216061-125-C22S-H',
            '01AcS-216061-125-C22S-H',  # 含小写的介质代码
            '01PL-216061A-125-C22S-H',  # 管道号带字母后缀
            '02WA-316052-200-D11T-N',
            '01PL-216061-80',
        ]
    elif project_type == "金昱元项目":
        patterns = [
            r'(\d{2}[A-Z]{1,4}\d{5,7}[A-Z]?)-(\d{2,4})-([A-Z0-9]{2,6})-([A-Z]{1,2})',
            r'(\d{2}[A-Z]{1,4}\d{5,7}[A-Z]?)-(\d{2,4})$',
        ]
        test_strings = [
            '01H04001-50-C1S-N',        # 1字母介质代码
            '01PC03012-100-C12S-H',     # 2字母介质代码
            '01CSL03203-150-C1N-H',     # 3字母介质代码
            '01IA04001A-50-C1S-N',      # 管道号带字母后缀
            '01PC03012-80',
        ]
    else:  # 巨化项目
        patterns = [
            r'(\d{4}[A-Z0-9]{1,4})-([A-Z0-9]{4,6})-(\d{2,4})-(\d{2}[A-Z0-9]{3,6})-([A-Z]{1,2})',
            r'(\d{4}[A-Z0-9]{1,4})-([A-Z0-9]{4,6})-(\d{2,4})$',
        ]
        test_strings = [
            '4101BRR-02457-200-03CBMB1-H',
            '4101BRR-02457-1000-03CBMB1-H',
            '4101CSM-01234-1200-02ABCD-H',
            '4101D-05678-50-01XYZ-C',
        ]

    for i, pattern in enumerate(patterns):
        log(f"测试模式 {i + 1}: {pattern}")
        for ts in test_strings:
            log(f"  测试字符串 '{ts}': {bool(re.search(pattern, ts))}")

    pipeline_numbers = []
    pattern_stats = {i: 0 for i in range(len(patterns))}

    log("开始分析前10个文本实体...")
    for idx, text in enumerate(text_entities[:10]):
        log(f"文本{idx}: {repr(text)} | 十六进制: {[hex(ord(c)) for c in str(text)[:20]]}")

    for text in text_entities:
        normalized = normalize_text(text)
        found = False
        for pidx, pattern in enumerate(patterns):
            for match in re.findall(pattern, normalized):
                pipeline_number = '-'.join(match) if isinstance(match, tuple) else match
                if pipeline_number not in pipeline_numbers:
                    pipeline_numbers.append(pipeline_number)
                    pattern_stats[pidx] += 1
                    log(f"找到管道号: {pipeline_number} (模式{pidx + 1}, 原文本: {repr(text[:50])})")
                    found = True
                    break
            if found:
                break

    log("各模式匹配统计:")
    for i, count in pattern_stats.items():
        log(f"  模式{i + 1}: {count}个匹配")

    return pipeline_numbers


def parse_pipeline_number(pipeline_number, medium_codes, project_type):
    """解析管道号，根据项目类型提取各字段"""
    parts = pipeline_number.split('-')

    if project_type == "金昱元项目":
        m = re.match(r'(\d{2})([A-Z]{1,4})(\d{5,7}[A-Z]?)', parts[0])
        if not m:
            return None
        unit_number, medium_code, pipe_number = m.group(1), m.group(2), m.group(3)
        if len(parts) >= 4:
            pipe_size, pipe_grade, insulation_grade = parts[1], parts[2], parts[3]
        elif len(parts) >= 2:
            pipe_size, pipe_grade, insulation_grade = parts[1], "未知等级", "未知"
        else:
            return None
        medium_name = medium_codes.get(medium_code, f"未知介质({medium_code})")
        return {
            'pipeline_number': parts[0],
            'unit_number': unit_number,
            'pipe_number': pipe_number,
            'nominal_diameter': pipe_size,
            'pipe_grade': pipe_grade,
            'insulation_grade': insulation_grade,
            'medium_code': medium_code,
            'medium_name': medium_name,
            'phase': determine_phase(medium_name),
        }

    elif project_type == "乌兹项目":
        if len(parts) >= 5:
            medium_code, pipe_number, pipe_size, pipe_grade, insulation_grade = parts[:5]
        elif len(parts) >= 3:
            medium_code, pipe_number, pipe_size = parts[:3]
            pipe_grade, insulation_grade = "未知等级", "未知"
        else:
            return None
        medium_name = medium_codes.get(medium_code, f"未知介质({medium_code})")
        return {
            'pipeline_number': f"{medium_code}-{pipe_number}",
            'unit_number': "",
            'pipe_number': pipe_number,
            'nominal_diameter': pipe_size,
            'pipe_grade': pipe_grade,
            'insulation_grade': insulation_grade,
            'medium_code': medium_code,
            'medium_name': medium_name,
            'phase': determine_phase(medium_name),
        }

    elif project_type == "天华项目":
        if len(parts) >= 5:
            unit_and_medium, pipe_number, pipe_size, pipe_grade, insulation_grade = parts[:5]
        elif len(parts) >= 3:
            unit_and_medium, pipe_number, pipe_size = parts[:3]
            pipe_grade, insulation_grade = "未知等级", "未知"
        else:
            return None
        m = re.match(r'(\d{2})([A-Za-z]{2,4})', unit_and_medium)
        unit_number = m.group(1) if m else unit_and_medium[:2]
        medium_code = m.group(2) if m else unit_and_medium[2:]
        medium_name = medium_codes.get(medium_code, f"未知介质({medium_code})")
        return {
            'pipeline_number': f"{unit_and_medium}-{pipe_number}",
            'unit_number': unit_number,
            'pipe_number': pipe_number,
            'nominal_diameter': pipe_size,
            'pipe_grade': pipe_grade,
            'insulation_grade': insulation_grade,
            'medium_code': medium_code,
            'medium_name': medium_name,
            'phase': determine_phase(medium_name),
        }

    else:  # 巨化项目
        if len(parts) >= 5:
            unit_and_medium, pipe_number, pipe_size, pipe_grade, insulation_grade = parts[:5]
        elif len(parts) >= 3:
            unit_and_medium, pipe_number, pipe_size = parts[:3]
            pipe_grade, insulation_grade = "未知等级", "未知"
        else:
            return None
        m = re.match(r'(\d{3,5})([A-Z0-9]+)', unit_and_medium)
        if m:
            unit_number, medium_code = m.group(1), m.group(2)
        else:
            unit_number = unit_and_medium[:4] if len(unit_and_medium) >= 4 else unit_and_medium
            medium_code = unit_and_medium[4:] if len(unit_and_medium) > 4 else ""
        medium_name = medium_codes.get(medium_code, f"未知介质({medium_code})")
        return {
            'pipeline_number': f"{unit_number}{medium_code}-{pipe_number}",
            'unit_number': unit_number,
            'pipe_number': pipe_number,
            'nominal_diameter': pipe_size,
            'pipe_grade': pipe_grade,
            'insulation_grade': insulation_grade,
            'medium_code': medium_code,
            'medium_name': medium_name,
            'phase': determine_phase(medium_name),
        }


def create_excel_output(pipeline_data, output_path):
    """将解析好的管道数据写入 Excel"""
    rows = [
        [d['pipeline_number'], d['nominal_diameter'], d['pipe_grade'],
         d['insulation_grade'], d['medium_name'], d['phase']]
        for d in pipeline_data if d
    ]
    columns = ['管道号', '管径', '管道等级', '保温等级', '介质名称', '相态']
    df = pd.DataFrame(rows, columns=columns)
    df = df.sort_values('管道号').reset_index(drop=True)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='管道数据表', index=False)
        ws = writer.sheets['管道数据表']
        for col, width in zip('ABCDEF', [20, 8, 15, 10, 15, 8]):
            ws.column_dimensions[col].width = width
        from openpyxl.styles import Font, PatternFill, Alignment
        font = Font(bold=True, color='FFFFFF')
        fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        align = Alignment(horizontal='center', vertical='center')
        for cell in ws[1]:
            cell.font = font
            cell.fill = fill
            cell.alignment = align

    logger.info(f"成功保存Excel文件: {output_path}")
    return df
